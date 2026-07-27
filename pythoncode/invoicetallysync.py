import os
import re
import sys
import json
import socket
import traceback
from functools import cmp_to_key
import mysql.connector
import requests
from datetime import datetime, date
import xml.sax.saxutils as saxutils
import xml.etree.ElementTree as ET
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from urllib.parse import urlparse
from urllib3.util import connection as urllib3_connection


def load_runtime_env():
    env_paths = [
        os.path.join(os.getcwd(), ".env"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"),
    ]
    for env_path in env_paths:
        if os.path.exists(env_path):
            load_dotenv(env_path, override=True)
            return
    load_dotenv(override=True)


load_runtime_env()


DB_CONFIG = {
    "host": "193.203.184.152",
    "user": "u380633007_lnpidata",
    "password": "!Office1@",
    "database": "u380633007_lnpidata",
    "port": 3306,
    "use_pure": True,
}

TALLY_URL = "http://127.0.0.1:9004"
TALLY_COMPANY_NAME = os.getenv("TALLY_COMPANY_NAME", "Laxmi Narayan Packaging Industries")
VOUCHER_TYPE_NAME = os.getenv("SALES_VOUCHER_TYPE_NAME", "Sales")
SALES_LEDGER_NAME = os.getenv("SALES_LEDGER_NAME", "Sales")
SALES_5_LEDGER_NAME = os.getenv("SALES_5_LEDGER_NAME", "SALES 5%")
SALES_18_LEDGER_NAME = os.getenv("SALES_18_LEDGER_NAME", "SALES 18%")
OTHER_CHARGES_LEDGER_NAME = os.getenv("SALES_OTHER_CHARGES_LEDGER_NAME", "Other Charges")
ROUND_OFF_LEDGER_NAME = os.getenv("SALES_ROUND_OFF_LEDGER_NAME", "Round Off")
CGST_LEDGER_NAME = os.getenv("OUTPUT_CGST_LEDGER_NAME", "Output CGST")
SGST_LEDGER_NAME = os.getenv("OUTPUT_SGST_LEDGER_NAME", "Output SGST")
IGST_LEDGER_NAME = os.getenv("OUTPUT_IGST_LEDGER_NAME", "Output IGST")
CGST_LEDGER_PREFIX = os.getenv("CGST_LEDGER_PREFIX", "Tax - CGST @")
SGST_LEDGER_PREFIX = os.getenv("SGST_LEDGER_PREFIX", "Tax - SGST @")
IGST_LEDGER_PREFIX = os.getenv("IGST_LEDGER_PREFIX", "Tax - IGST @")
DEBUG_TALLY_XML = os.getenv("DEBUG_TALLY_XML", "0").strip() == "1"
DEFAULT_UPDATED_BY = os.getenv("TALLY_SYNC_USER", "system")
TALLY_UOM_ALIASES = {
    "KG": os.getenv("TALLY_UOM_KG", "KG"),
    "KGS": os.getenv("TALLY_UOM_KGS", "KG"),
    "KILOGRAM": os.getenv("TALLY_UOM_KILOGRAM", "KG"),
    "KILOGRAMS": os.getenv("TALLY_UOM_KILOGRAMS", "KG"),
    "NOS": os.getenv("TALLY_UOM_NOS", "NOS"),
    "NO": os.getenv("TALLY_UOM_NO", "NOS"),
    "PCS": os.getenv("TALLY_UOM_PCS", "PCS"),
    "PC": os.getenv("TALLY_UOM_PC", "PCS"),
}
TALLY_MASTER_CACHE = {}
TALLY_CONNECT_TIMEOUT = float(os.getenv("TALLY_CONNECT_TIMEOUT", "5"))
TALLY_READ_TIMEOUT = float(os.getenv("TALLY_READ_TIMEOUT", "60"))
LOG_FOLDER_NAME = "Log"
LOG_WORKBOOK_NAME = "invoice_sync_logs.xlsx"
LOG_SHEET_NAME = "SyncLogs"
TALLY_DEBUG_FOLDER_NAME = "tally_debug"
LOG_HEADERS = [
    "Timestamp",
    "Status",
    "Stage",
    "Invoice ID",
    "Invoice No",
    "Company Name",
    "Message",
    "Tally Invoice No",
    "Tally Invoice Date",
    "Tally Invoice ID",
    "Runtime Path",
]
TALLY_SYNC_API_URL = os.getenv(
    "TALLY_SYNC_API_URL",
    "https://darkred-lobster-409686.hostingersite.com",
).strip().rstrip("/")
TALLY_SYNC_API_SECRET = os.getenv("TALLY_SYNC_API_SECRET", "!Office1@").strip()
FORCE_IPV4_HTTP = os.getenv("TALLY_SYNC_FORCE_IPV4", "1").strip() != "0"
_HTTP_IPV4_PATCHED = False



def log_terminal(level, message):
    print(f"[{level}] {message}")


def use_tally_sync_api():
    return bool(TALLY_SYNC_API_URL and TALLY_SYNC_API_SECRET)


def ensure_ipv4_http():
    global _HTTP_IPV4_PATCHED
    if _HTTP_IPV4_PATCHED or not FORCE_IPV4_HTTP:
        return

    # Prefer IPv4 for packaged builds to avoid Windows socket-policy issues
    # seen on some machines when urllib/requests chooses IPv6 first.
    urllib3_connection.allowed_gai_family = lambda: socket.AF_INET
    _HTTP_IPV4_PATCHED = True


def build_connectivity_error(url, exc):
    parsed = urlparse(url)
    host = parsed.hostname or url
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    message = str(exc)

    if "WinError 10013" in message:
        return (
            f"Could not open HTTPS connection to {host}:{port}. "
            "Windows is refusing socket access for this EXE process "
            "(WinError 10013). "
            "Please allow this EXE in Windows Security/antivirus, or run "
            "the Python script directly if Python is already able to access "
            "the same URL."
        )

    return f"Could not connect to {host}:{port}. Details: {message}"


def call_tally_sync_api(method, path, payload=None, params=None):
    if not use_tally_sync_api():
        raise RuntimeError("Tally Sync API mode is not configured")

    ensure_ipv4_http()
    url = f"{TALLY_SYNC_API_URL}{path}"
    try:
        response = requests.request(
            method=method.upper(),
            url=url,
            json=payload,
            params=params,
            headers={
                "x-tally-sync-secret": TALLY_SYNC_API_SECRET,
                "Content-Type": "application/json",
            },
            timeout=(TALLY_CONNECT_TIMEOUT, max(TALLY_READ_TIMEOUT, 30)),
        )
    except requests.exceptions.RequestException as exc:
        raise RuntimeError(build_connectivity_error(url, exc)) from exc

    try:
        body = response.json()
    except Exception:
        body = {"error": response.text[:1000]}

    if response.status_code >= 400:
        raise RuntimeError(body.get("error") or f"HTTP {response.status_code}")

    return body


def get_pending_invoice_rows_api():
    rows = call_tally_sync_api("GET", "/api/tally-sync/pending-invoices")
    return sorted(rows or [], key=cmp_to_key(compare_pending_invoice_rows))


def get_invoice_context_api(invoice_id):
    return call_tally_sync_api("GET", f"/api/tally-sync/invoices/{invoice_id}/context")


def update_invoice_tally_status_api(
    invoice_id,
    success,
    remark,
    tally_by=None,
    tally_inv_no=None,
    tally_inv_date=None,
    tally_inv_id=None,
):
    call_tally_sync_api(
        "POST",
        f"/api/tally-sync/invoices/{invoice_id}/status",
        payload={
            "success": bool(success),
            "remark": remark,
            "tallyBy": tally_by or DEFAULT_UPDATED_BY,
            "tallyInvNo": tally_inv_no,
            "tallyInvDate": tally_inv_date,
            "tallyInvId": tally_inv_id,
        },
    )


def get_runtime_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def get_log_workbook_path():
    log_dir = os.path.join(get_runtime_base_dir(), LOG_FOLDER_NAME)
    os.makedirs(log_dir, exist_ok=True)
    return os.path.join(log_dir, LOG_WORKBOOK_NAME)


def pause_before_exit():
    if not getattr(sys, "frozen", False):
        return
    try:
        input("Press Enter to close...")
    except EOFError:
        pass


def get_log_dir():
    log_dir = os.path.join(get_runtime_base_dir(), LOG_FOLDER_NAME)
    os.makedirs(log_dir, exist_ok=True)
    return log_dir


def write_tally_debug_dump(prefix, item_name, xml_text):
    try:
        debug_dir = os.path.join(get_log_dir(), TALLY_DEBUG_FOLDER_NAME)
        os.makedirs(debug_dir, exist_ok=True)
        safe_name = re.sub(r'[^A-Za-z0-9._-]+', "_", str(item_name or "").strip())[:120] or "unknown_item"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(debug_dir, f"{prefix}_{timestamp}_{safe_name}.xml")
        with open(file_path, "w", encoding="utf-8") as handle:
            handle.write(xml_text or "")
        return file_path
    except Exception as debug_error:
        log_terminal("LOG", f"Could not write Tally debug XML: {debug_error}")
        return ""


def ensure_log_workbook():
    workbook_path = get_log_workbook_path()

    if os.path.exists(workbook_path):
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()

    if LOG_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[LOG_SHEET_NAME]
    else:
        if workbook.active and workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None:
            sheet = workbook.active
            sheet.title = LOG_SHEET_NAME
        else:
            sheet = workbook.create_sheet(LOG_SHEET_NAME)
        sheet.append(LOG_HEADERS)

    if sheet.max_row == 1:
        current_headers = [sheet.cell(row=1, column=index + 1).value for index in range(len(LOG_HEADERS))]
        if current_headers != LOG_HEADERS:
            for index, header in enumerate(LOG_HEADERS, start=1):
                sheet.cell(row=1, column=index).value = header

    workbook.save(workbook_path)
    return workbook_path


def append_excel_log(
    status,
    stage,
    message,
    invoice_row=None,
    company_name="",
    tally_inv_no="",
    tally_inv_date="",
    tally_inv_id="",
):
    try:
        workbook_path = ensure_log_workbook()
        workbook = load_workbook(workbook_path)
        sheet = workbook[LOG_SHEET_NAME]
        runtime_path = sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        sheet.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(status or "").strip(),
                str(stage or "").strip(),
                str((invoice_row or {}).get("id") or "").strip(),
                str((invoice_row or {}).get("invoiceNo") or "").strip(),
                str(company_name or "").strip(),
                str(message or "").strip(),
                str(tally_inv_no or "").strip(),
                str(tally_inv_date or "").strip(),
                str(tally_inv_id or "").strip(),
                runtime_path,
            ]
        )
        workbook.save(workbook_path)
        workbook.close()
    except Exception as log_error:
        log_terminal("LOG", f"Could not write Excel log: {log_error}")


def esc(value):
    if value is None:
        return ""
    return saxutils.escape(str(value))


def to_float(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def format_tally_date(value):
    if value is None or value == "":
        return datetime.today().strftime("%Y%m%d")

    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")

    if isinstance(value, date):
        return value.strftime("%Y%m%d")

    raw_value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw_value, fmt).strftime("%Y%m%d")
        except Exception:
            pass

    return datetime.today().strftime("%Y%m%d")


def normalize_uom(value):
    if value is None:
        return ""

    normalized = str(value).strip()
    if not normalized:
        return ""

    return TALLY_UOM_ALIASES.get(normalized.upper(), normalized)


def normalize_part_no(value):
    if value is None:
        return ""

    normalized = re.sub(r"\s+", " ", str(value).strip())
    return normalized.upper()


def join_unique_values(values):
    unique_values = []
    seen = set()

    for value in values or []:
        normalized = str(value or "").strip()
        if not normalized:
            continue
        key = normalized.upper()
        if key in seen:
            continue
        seen.add(key)
        unique_values.append(normalized)

    return ", ".join(unique_values)


def sanitize_tally_xml(xml_text):
    if not xml_text:
        return xml_text
    cleaned = re.sub(r"&#x0*([0-8BCEF]|1[0-9A-F]);", "", xml_text, flags=re.IGNORECASE)
    cleaned = re.sub(r"&#([0-8]|1[0-9]|2[0-9]|30|31);", "", cleaned, flags=re.IGNORECASE)
    return cleaned


def format_iso_date(value):
    if value is None or value == "":
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    raw_value = str(value).strip()
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw_value, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    return raw_value


def normalize_invoice_number(value):
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def is_probable_voucher_element(element):
    if element is None:
        return False
    if element.attrib.get("VOUCHERNUMBER") or element.attrib.get("GUID") or element.attrib.get("VOUCHERKEY"):
        return True
    child_tags = {str(child.tag or "").upper() for child in list(element)}
    return bool(child_tags.intersection({"VOUCHERNUMBER", "DATE", "GUID", "VOUCHERKEY", "REMOTEID", "PARTYLEDGERNAME"}))


def find_first_voucher_element(root):
    for candidate in root.findall(".//VOUCHER"):
        if is_probable_voucher_element(candidate):
            return candidate
    return None


def parse_tally_voucher_response(response_text):
    if not response_text:
        return {}

    cleaned = sanitize_tally_xml(response_text)
    result = {}

    try:
        root = ET.fromstring(cleaned)
        voucher = find_first_voucher_element(root)
        if voucher is not None:
            result.update(extract_voucher_summary(voucher))
        else:
            # Fallback search in entire response for common identifiers
            if "VOUCHERNUMBER" in cleaned.upper():
                match = re.search(r"<VOUCHERNUMBER>([^<]+)</VOUCHERNUMBER>", cleaned, re.IGNORECASE)
                if match:
                    result["tallyInvNo"] = match.group(1).strip()
            if "<DATE>" in cleaned.upper():
                match = re.search(r"<DATE>([^<]+)</DATE>", cleaned, re.IGNORECASE)
                if match:
                    result["tallyInvDate"] = match.group(1).strip()
            for key in ("GUID", "VOUCHERKEY", "REMOTEID"):
                pattern = rf"<{key}>([^<]+)</{key}>"
                match = re.search(pattern, cleaned, re.IGNORECASE)
                if match:
                    result["tallyInvId"] = match.group(1).strip()
                    break
            for key, output_key in (("LASTVCHID", "lastVchId"), ("LASTMID", "lastMasterId")):
                pattern = rf"<{key}>([^<]+)</{key}>"
                match = re.search(pattern, cleaned, re.IGNORECASE)
                if match:
                    result[output_key] = match.group(1).strip()
    except Exception:
        pass

    return {k: v for k, v in result.items() if v}


def extract_voucher_summary(voucher):
    summary = {
        "tallyInvNo": voucher.get("VOUCHERNUMBER") or "",
        "tallyInvDate": voucher.get("DATE") or "",
        "tallyInvId": voucher.get("GUID") or voucher.get("VOUCHERKEY") or voucher.get("REMOTEID") or "",
        "partyLedgerName": voucher.get("PARTYLEDGERNAME") or "",
        "voucherTypeName": voucher.get("VOUCHERTYPENAME") or "",
        "narration": "",
    }

    for child in voucher:
        tag = child.tag.upper()
        text = (child.text or "").strip()
        if not text:
            continue
        if tag == "VOUCHERNUMBER" and not summary["tallyInvNo"]:
            summary["tallyInvNo"] = text
        elif tag == "DATE" and not summary["tallyInvDate"]:
            summary["tallyInvDate"] = text
        elif tag in {"GUID", "VOUCHERKEY", "REMOTEID"} and not summary["tallyInvId"]:
            summary["tallyInvId"] = text
        elif tag == "PARTYLEDGERNAME" and not summary["partyLedgerName"]:
            summary["partyLedgerName"] = text
        elif tag == "VOUCHERTYPENAME" and not summary["voucherTypeName"]:
            summary["voucherTypeName"] = text
        elif tag == "NARRATION" and not summary["narration"]:
            summary["narration"] = text

    return {key: value for key, value in summary.items() if value}


def parse_tally_voucher_collection(response_text):
    if not response_text:
        return []

    cleaned = sanitize_tally_xml(response_text)
    try:
        root = ET.fromstring(cleaned)
    except Exception:
        return []

    vouchers = []
    for voucher in root.findall(".//VOUCHER"):
        if not is_probable_voucher_element(voucher):
            continue
        summary = extract_voucher_summary(voucher)
        if summary:
            vouchers.append(summary)
    return vouchers


def fetch_tally_vouchers_for_date(invoice_date, voucher_type=None):
    tally_date = format_tally_date(invoice_date)
    if not tally_date:
        return []

    xml = f"""
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <EXPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Voucher Register</REPORTNAME>
                <STATICVARIABLES>
                    <SVFROMDATE>{esc(tally_date)}</SVFROMDATE>
                    <SVTODATE>{esc(tally_date)}</SVTODATE>
                    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                </STATICVARIABLES>
                <FETCHLIST>
                    <FETCH>Date</FETCH>
                    <FETCH>VoucherNumber</FETCH>
                    <FETCH>GUID</FETCH>
                    <FETCH>VoucherKey</FETCH>
                    <FETCH>RemoteID</FETCH>
                    <FETCH>PartyLedgerName</FETCH>
                    <FETCH>VoucherTypeName</FETCH>
                    <FETCH>Narration</FETCH>
                </FETCHLIST>
            </REQUESTDESC>
        </EXPORTDATA>
    </BODY>
</ENVELOPE>
"""
    response_text = tally_request(xml)
    vouchers = parse_tally_voucher_collection(response_text)
    if not voucher_type:
        return vouchers

    normalized_voucher_type = str(voucher_type or "").strip().upper()
    return [
        voucher
        for voucher in vouchers
        if str(voucher.get("voucherTypeName") or "").strip().upper() == normalized_voucher_type
    ]


def fetch_tally_voucher_reference(invoice_no, voucher_type=None, invoice_date=None, party_name=None):
    if not invoice_no:
        return {}

    normalized_invoice_no = normalize_invoice_number(invoice_no)
    normalized_party_name = str(party_name or "").strip().upper()

    if invoice_date:
        vouchers = fetch_tally_vouchers_for_date(invoice_date, voucher_type=voucher_type)
        exact_matches = [
            voucher
            for voucher in vouchers
            if normalize_invoice_number(voucher.get("tallyInvNo")) == normalized_invoice_no
        ]
        if normalized_party_name:
            party_matches = [
                voucher
                for voucher in exact_matches
                if str(voucher.get("partyLedgerName") or "").strip().upper() == normalized_party_name
            ]
            if party_matches:
                return party_matches[-1]
        if exact_matches:
            return exact_matches[-1]

    voucher_type_attr = f' VOUCHERTYPENAME="{esc(voucher_type)}"' if voucher_type else ""
    xml = f"""
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <EXPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Voucher Register</REPORTNAME>
                <STATICVARIABLES>
                    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                </STATICVARIABLES>
            </REQUESTDESC>
            <REQUESTDATA>
                <TALLYMESSAGE>
                    <VOUCHER VOUCHERNUMBER="{esc(invoice_no)}"{voucher_type_attr} ACTION="Get" />
                </TALLYMESSAGE>
            </REQUESTDATA>
        </EXPORTDATA>
    </BODY>
</ENVELOPE>
"""
    response_text = tally_request(xml)
    voucher = parse_tally_voucher_response(response_text)
    if normalize_invoice_number(voucher.get("tallyInvNo")) == normalized_invoice_no:
        return voucher
    return {}


def fetch_tally_voucher_by_id(tally_inv_id):
    if not tally_inv_id:
        return {}

    object_request_variants = ("GUID", "MasterID", "VoucherKey")
    for id_type in object_request_variants:
        xml = f"""
<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Object</TYPE>
        <SUBTYPE>Voucher</SUBTYPE>
        <ID TYPE="{esc(id_type)}">{esc(tally_inv_id)}</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <FETCHLIST>
                <FETCH>Date</FETCH>
                <FETCH>VoucherNumber</FETCH>
                <FETCH>GUID</FETCH>
                <FETCH>VoucherKey</FETCH>
                <FETCH>RemoteID</FETCH>
            </FETCHLIST>
        </DESC>
    </BODY>
</ENVELOPE>
"""
        response_text = tally_request(xml)
        result = parse_tally_voucher_response(response_text)
        if result.get("tallyInvId") or result.get("tallyInvNo"):
            return result

    action_get_variants = (
        f'<VOUCHER MASTERID="{esc(tally_inv_id)}" ACTION="Get"></VOUCHER>',
        f'<VOUCHER GUID="{esc(tally_inv_id)}" ACTION="Get"></VOUCHER>',
        f'<VOUCHER VCHKEY="{esc(tally_inv_id)}" ACTION="Get"></VOUCHER>',
        f'<VOUCHER REMOTEID="{esc(tally_inv_id)}" ACTION="Get"></VOUCHER>',
    )

    for voucher_tag in action_get_variants:
        xml = f"""
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <EXPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Voucher Register</REPORTNAME>
                <STATICVARIABLES>
                    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                </STATICVARIABLES>
            </REQUESTDESC>
            <REQUESTDATA>
                <TALLYMESSAGE>
                    {voucher_tag}
                </TALLYMESSAGE>
            </REQUESTDATA>
        </EXPORTDATA>
    </BODY>
</ENVELOPE>
"""
        response_text = tally_request(xml)
        result = parse_tally_voucher_response(response_text)
        if result.get("tallyInvId") or result.get("tallyInvNo"):
            return result

    return {}


def fetch_created_tally_voucher(response_text):
    parsed_response = parse_tally_voucher_response(response_text)
    for candidate in (
        parsed_response.get("tallyInvId"),
        parsed_response.get("lastVchId"),
        parsed_response.get("lastMasterId"),
    ):
        if not candidate:
            continue
        voucher = fetch_tally_voucher_by_id(candidate)
        if voucher:
            if not voucher.get("tallyInvId"):
                voucher["tallyInvId"] = candidate
            return voucher
    if parsed_response.get("lastVchId") and not parsed_response.get("tallyInvId"):
        parsed_response["tallyInvId"] = parsed_response.get("lastVchId")
    elif parsed_response.get("lastMasterId") and not parsed_response.get("tallyInvId"):
        parsed_response["tallyInvId"] = parsed_response.get("lastMasterId")
    return parsed_response


def fetch_tally_voucher_by_context(invoice_row, customer_name, narration_text):
    tally_date = format_tally_date(invoice_row.get("date"))
    xml = f"""
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <EXPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Voucher Register</REPORTNAME>
                <STATICVARIABLES>
                    <SVFROMDATE>{esc(tally_date)}</SVFROMDATE>
                    <SVTODATE>{esc(tally_date)}</SVTODATE>
                    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                </STATICVARIABLES>
                <FETCHLIST>
                    <FETCH>Date</FETCH>
                    <FETCH>VoucherNumber</FETCH>
                    <FETCH>GUID</FETCH>
                    <FETCH>VoucherKey</FETCH>
                    <FETCH>RemoteID</FETCH>
                    <FETCH>PartyLedgerName</FETCH>
                    <FETCH>Narration</FETCH>
                </FETCHLIST>
            </REQUESTDESC>
        </EXPORTDATA>
    </BODY>
</ENVELOPE>
"""
    response_text = tally_request(xml)
    vouchers = parse_tally_voucher_collection(response_text)
    if not vouchers:
        return {}

    normalized_date = format_iso_date(tally_date)
    normalized_party = str(customer_name or "").strip().upper()
    normalized_narration = str(narration_text or "").strip().upper()

    exact_matches = [
        voucher
        for voucher in vouchers
        if format_iso_date(voucher.get("tallyInvDate")) == normalized_date
        and str(voucher.get("partyLedgerName") or "").strip().upper() == normalized_party
        and str(voucher.get("narration") or "").strip().upper() == normalized_narration
    ]
    if exact_matches:
        return exact_matches[-1]

    party_matches = [
        voucher
        for voucher in vouchers
        if format_iso_date(voucher.get("tallyInvDate")) == normalized_date
        and str(voucher.get("partyLedgerName") or "").strip().upper() == normalized_party
    ]
    if party_matches:
        return party_matches[-1]

    return {}


def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)


def get_db_cursor(conn, dictionary=False):
    try:
        if conn and hasattr(conn, "ping"):
            conn.ping(reconnect=True, attempts=3, delay=2)
    except Exception:
        pass
    return conn.cursor(dictionary=dictionary)


def column_exists(conn, table_name, column_name):
    cursor = get_db_cursor(conn)
    cursor.execute(
        """
        SELECT 1
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = %s
          AND TABLE_NAME = %s
          AND COLUMN_NAME = %s
        LIMIT 1
        """,
        (DB_CONFIG["database"], table_name, column_name),
    )
    exists = cursor.fetchone() is not None
    cursor.close()
    return exists


def ensure_invoice_sync_columns(conn):
    cursor = get_db_cursor(conn)
    try:
        if not column_exists(conn, "invoices", "tallySyncRemark"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallySyncRemark TEXT")
        if not column_exists(conn, "invoices", "tallyBy"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallyBy VARCHAR(255)")
        if not column_exists(conn, "invoices", "tallyTimestamp"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallyTimestamp VARCHAR(255)")
        if not column_exists(conn, "invoices", "tallyInvNo"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallyInvNo VARCHAR(255)")
        if not column_exists(conn, "invoices", "tallyInvDate"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallyInvDate VARCHAR(255)")
        if not column_exists(conn, "invoices", "tallyInvId"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN tallyInvId VARCHAR(255)")
        if not column_exists(conn, "invoices", "otherChargesGstRate"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN otherChargesGstRate DECIMAL(5,2) NULL DEFAULT NULL")
        if not column_exists(conn, "invoices", "otherChargesCgst"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN otherChargesCgst DECIMAL(15,2) NOT NULL DEFAULT 0")
        if not column_exists(conn, "invoices", "otherChargesSgst"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN otherChargesSgst DECIMAL(15,2) NOT NULL DEFAULT 0")
        if not column_exists(conn, "invoices", "otherChargesIgst"):
            cursor.execute("ALTER TABLE invoices ADD COLUMN otherChargesIgst DECIMAL(15,2) NOT NULL DEFAULT 0")
        conn.commit()
    finally:
        cursor.close()


def parse_invoice_no_parts(invoice_no):
    raw = str(invoice_no or "").strip()
    if not raw:
        return None

    parts = [part.strip() for part in raw.split("/")]
    if len(parts) < 3:
        return None

    prefix = "/".join(parts[:-2]).strip()
    fy_label = parts[-2]
    suffix = parts[-1]
    fy_match = re.fullmatch(r"(\d{2,4})-(\d{2,4})", fy_label)
    if not fy_match:
        return None

    start_year_text = fy_match.group(1)
    start_year = int(start_year_text)
    if len(start_year_text) == 2:
        start_year += 2000

    if not suffix.isdigit():
        return None

    return {
        "prefix": prefix,
        "fy_label": fy_label,
        "fy_start_year": start_year,
        "suffix_number": int(suffix),
    }


def compare_pending_invoice_rows(left_row, right_row):
    left_invoice_no = str(left_row.get("invoiceNo") or "").strip()
    right_invoice_no = str(right_row.get("invoiceNo") or "").strip()
    left_parts = parse_invoice_no_parts(left_invoice_no)
    right_parts = parse_invoice_no_parts(right_invoice_no)

    if left_parts and right_parts:
        left_key = (
            left_parts["fy_start_year"],
            left_parts["suffix_number"],
            left_invoice_no,
            str(left_row.get("date") or ""),
            str(left_row.get("id") or ""),
        )
        right_key = (
            right_parts["fy_start_year"],
            right_parts["suffix_number"],
            right_invoice_no,
            str(right_row.get("date") or ""),
            str(right_row.get("id") or ""),
        )
        if left_key < right_key:
            return -1
        if left_key > right_key:
            return 1
        return 0

    if left_parts and not right_parts:
        return -1
    if right_parts and not left_parts:
        return 1

    left_fallback = (
        str(left_row.get("date") or ""),
        left_invoice_no,
        str(left_row.get("id") or ""),
    )
    right_fallback = (
        str(right_row.get("date") or ""),
        right_invoice_no,
        str(right_row.get("id") or ""),
    )
    if left_fallback < right_fallback:
        return -1
    if left_fallback > right_fallback:
        return 1
    return 0


def get_pending_invoice_rows(conn):
    sql = """
        SELECT *
        FROM invoices
        WHERE tallyTimestamp IS NULL
           OR tallyTimestamp = ''
    """

    cursor = get_db_cursor(conn, dictionary=True)
    cursor.execute(sql)
    rows = cursor.fetchall()
    cursor.close()
    return sorted(rows, key=cmp_to_key(compare_pending_invoice_rows))


def tally_request(xml_data):
    try:
        response = requests.post(
            TALLY_URL,
            data=xml_data.encode("utf-8"),
            headers={"Content-Type": "text/xml"},
            timeout=(TALLY_CONNECT_TIMEOUT, TALLY_READ_TIMEOUT),
        )
        if response.status_code == 200:
            return response.text
        return f"HTTP error from Tally: {response.status_code}"
    except requests.exceptions.ReadTimeout:
        return f"Connection error: Tally request timed out after {TALLY_READ_TIMEOUT} seconds while waiting for response"
    except requests.exceptions.ConnectTimeout:
        return f"Connection error: Could not connect to Tally within {TALLY_CONNECT_TIMEOUT} seconds"
    except Exception as error:
        return f"Connection error: {error}"


def check_tally_object_exists(object_type, object_name):
    if not object_name:
        return False, "Empty name"

    cache_key = (object_type.upper(), object_name.strip().upper())
    if cache_key in TALLY_MASTER_CACHE:
        return TALLY_MASTER_CACHE[cache_key]

    xml = f"""<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Object</TYPE>
        <SUBTYPE>{esc(object_type)}</SUBTYPE>
        <ID TYPE="Name">{esc(object_name)}</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <FETCHLIST>
                <FETCH>Name</FETCH>
            </FETCHLIST>
        </DESC>
    </BODY>
</ENVELOPE>"""

    response_text = tally_request(xml)
    if not response_text:
        result = (False, "Empty response from Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    upper_response = response_text.upper()
    if "CONNECTION ERROR:" in upper_response or "HTTP ERROR FROM TALLY:" in upper_response:
        result = (False, response_text)
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    if "UNKNOWN REQUEST" in upper_response or "<LINEERROR>" in upper_response:
        result = (False, f"{object_type} '{object_name}' not found in Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    try:
        root = ET.fromstring(sanitize_tally_xml(response_text))
        found = root.find(f".//{object_type.upper().replace(' ', '')}")
        if found is not None or object_name.upper() in upper_response:
            result = (True, "")
        else:
            result = (False, f"{object_type} '{object_name}' not found in Tally")
    except Exception:
        if object_name.upper() in upper_response:
            result = (True, "")
        else:
            result = (False, f"{object_type} '{object_name}' not found in Tally")

    TALLY_MASTER_CACHE[cache_key] = result
    return result


def extract_first_matching_tag(xml_block, tag_names):
    for tag_name in tag_names:
        escaped_tag = re.escape(tag_name)
        pattern = rf"<{escaped_tag}\b[^>]*>(.*?)</{escaped_tag}>"
        match = re.search(pattern, xml_block, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return re.sub(r"\s+", " ", match.group(1).strip())
    return ""


def extract_recursive_tag_text(element, tag_names):
    normalized_tags = {str(tag or "").upper() for tag in tag_names}
    for node in element.iter():
        tag_name = str(node.tag or "").upper()
        if tag_name in normalized_tags:
            text_value = re.sub(r"\s+", " ", "".join(node.itertext()).strip())
            if text_value:
                return text_value
    return ""


def fetch_tally_stock_item_details(item_name):
    if not item_name:
        return {}, "Stock Item name missing"

    cache_key = ("STOCK_ITEM_DETAILS", item_name.strip().upper())
    if cache_key in TALLY_MASTER_CACHE:
        return TALLY_MASTER_CACHE[cache_key]

    xml = f"""<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Object</TYPE>
        <SUBTYPE>Stock Item</SUBTYPE>
        <ID TYPE="Name">{esc(item_name)}</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <FETCHLIST>
                <FETCH>Name</FETCH>
                <FETCH>BaseUnits</FETCH>
                <FETCH>PartNo</FETCH>
                <FETCH>PartNumber</FETCH>
                <FETCH>MailingName.LIST</FETCH>
                <FETCH>LanguageName.LIST</FETCH>
            </FETCHLIST>
        </DESC>
    </BODY>
</ENVELOPE>"""

    response_text = tally_request(xml)
    if not response_text:
        result = ({}, "Empty response from Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    upper_response = response_text.upper()
    if "CONNECTION ERROR:" in upper_response or "HTTP ERROR FROM TALLY:" in upper_response:
        result = ({}, response_text)
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    if "UNKNOWN REQUEST" in upper_response or "<LINEERROR>" in upper_response:
        result = ({}, f"Stock Item '{item_name}' not found in Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    cleaned_xml = sanitize_tally_xml(response_text)
    try:
        root = ET.fromstring(cleaned_xml)
    except Exception:
        result = ({}, f"Could not read Stock Item details for '{item_name}' from Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    stock_item = None
    for candidate in root.findall(".//STOCKITEM"):
        if candidate.attrib.get("NAME") or candidate.attrib.get("REQNAME") or candidate.attrib.get("ID"):
            stock_item = candidate
            break
    if stock_item is None:
        result = ({}, f"Stock Item '{item_name}' not found in Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    try:
        stock_item_block = ET.tostring(stock_item, encoding="unicode")
    except Exception:
        stock_item_block = cleaned_xml

    base_units = extract_recursive_tag_text(
        stock_item,
        ["BASEUNITS", "BASEUNIT", "BASEUNAME", "UNITS", "UNIT"],
    )
    if not base_units:
        base_units = extract_first_matching_tag(
            stock_item_block,
            ["BASEUNITS", "BASEUNIT", "BASEUNAME", "UNITS", "UNIT"],
        )

    normalized_base_units = normalize_uom(base_units)
    if not normalized_base_units:
        debug_path = write_tally_debug_dump("stock_item_no_unit", item_name, cleaned_xml)
        debug_suffix = f" Debug XML: {debug_path}" if debug_path else ""
        result = ({}, f"Stock Item '{item_name}' has no Base Unit in Tally (or parser could not read it).{debug_suffix}")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    part_no = extract_first_matching_tag(stock_item_block, ["PARTNO", "PARTNUMBER"])

    mailing_names = re.findall(
        r"<MAILINGNAME\b[^>]*>(.*?)</MAILINGNAME>",
        stock_item_block,
        flags=re.IGNORECASE | re.DOTALL,
    )
    language_names = re.findall(
        r"<NAME\b[^>]*>(.*?)</NAME>",
        stock_item_block,
        flags=re.IGNORECASE | re.DOTALL,
    )

    result = (
        {
            "base_uom": normalized_base_units,
            "part_no": re.sub(r"\s+", " ", part_no.strip()) if part_no else "",
            "mailing_names": [re.sub(r"\s+", " ", value.strip()) for value in mailing_names if value.strip()],
            "language_names": [re.sub(r"\s+", " ", value.strip()) for value in language_names if value.strip()],
        },
        "",
    )
    TALLY_MASTER_CACHE[cache_key] = result
    return result


def fetch_tally_ledger_details(ledger_name):
    if not ledger_name:
        return {}, "Ledger name missing"

    cache_key = ("LEDGER_DETAILS", ledger_name.strip().upper())
    if cache_key in TALLY_MASTER_CACHE:
        return TALLY_MASTER_CACHE[cache_key]

    xml = f"""<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Object</TYPE>
        <SUBTYPE>Ledger</SUBTYPE>
        <ID TYPE="Name">{esc(ledger_name)}</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <FETCHLIST>
                <FETCH>Name</FETCH>
                <FETCH>MailingName</FETCH>
                <FETCH>MailingName.LIST</FETCH>
                <FETCH>Address</FETCH>
                <FETCH>Address.LIST</FETCH>
                <FETCH>StateName</FETCH>
                <FETCH>CountryName</FETCH>
                <FETCH>PINCode</FETCH>
                <FETCH>IncomeTaxNumber</FETCH>
                <FETCH>GSTRegistrationType</FETCH>
                <FETCH>PartyGSTIN</FETCH>
                <FETCH>GSTIN</FETCH>
            </FETCHLIST>
        </DESC>
    </BODY>
</ENVELOPE>"""

    response_text = tally_request(xml)
    if not response_text:
        result = ({}, "Empty response from Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    upper_response = response_text.upper()
    if "CONNECTION ERROR:" in upper_response or "HTTP ERROR FROM TALLY:" in upper_response:
        result = ({}, response_text)
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    if "UNKNOWN REQUEST" in upper_response or "<LINEERROR>" in upper_response:
        result = ({}, f"Ledger '{ledger_name}' not found in Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    cleaned_xml = sanitize_tally_xml(response_text)
    try:
        root = ET.fromstring(cleaned_xml)
    except Exception:
        result = ({}, f"Could not read Ledger details for '{ledger_name}' from Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    ledger = None
    for candidate in root.findall(".//LEDGER"):
        if candidate.attrib.get("NAME") or candidate.attrib.get("REQNAME") or candidate.attrib.get("ID"):
            ledger = candidate
            break
    if ledger is None:
        result = ({}, f"Ledger '{ledger_name}' not found in Tally")
        TALLY_MASTER_CACHE[cache_key] = result
        return result

    try:
        ledger_block = ET.tostring(ledger, encoding="unicode")
    except Exception:
        ledger_block = cleaned_xml

    mailing_name = extract_recursive_tag_text(ledger, ["MAILINGNAME", "NAME"]) or extract_first_matching_tag(
        ledger_block,
        ["MAILINGNAME", "NAME"],
    )
    state_name = extract_recursive_tag_text(ledger, ["STATENAME"]) or extract_first_matching_tag(ledger_block, ["STATENAME"])
    country_name = extract_recursive_tag_text(ledger, ["COUNTRYNAME"]) or extract_first_matching_tag(ledger_block, ["COUNTRYNAME"])
    pin_code = extract_recursive_tag_text(ledger, ["PINCODE"]) or extract_first_matching_tag(ledger_block, ["PINCODE"])
    gst_registration_type = extract_recursive_tag_text(
        ledger,
        ["GSTREGISTRATIONTYPE", "REGISTRATIONTYPE"],
    ) or extract_first_matching_tag(ledger_block, ["GSTREGISTRATIONTYPE", "REGISTRATIONTYPE"])
    gstin = extract_recursive_tag_text(ledger, ["PARTYGSTIN", "GSTIN"]) or extract_first_matching_tag(
        ledger_block,
        ["PARTYGSTIN", "GSTIN"],
    )

    address_lines = re.findall(
        r"<ADDRESS\b[^>]*>(.*?)</ADDRESS>",
        ledger_block,
        flags=re.IGNORECASE | re.DOTALL,
    )
    normalized_address_lines = [
        sanitize_party_address_line(value)
        for value in address_lines
        if sanitize_party_address_line(value)
    ]

    result = (
        {
            "name": str(ledger.attrib.get("NAME") or ledger_name).strip(),
            "mailing_name": re.sub(r"\s+", " ", mailing_name.strip()) if mailing_name else "",
            "address_lines": normalized_address_lines,
            "state": re.sub(r"\s+", " ", state_name.strip()) if state_name else "",
            "country": re.sub(r"\s+", " ", country_name.strip()) if country_name else "",
            "pin": re.sub(r"\s+", " ", pin_code.strip()) if pin_code else "",
            "gst_registration_type": re.sub(r"\s+", " ", gst_registration_type.strip()) if gst_registration_type else "",
            "gstin": re.sub(r"\s+", " ", gstin.strip()) if gstin else "",
        },
        "",
    )
    TALLY_MASTER_CACHE[cache_key] = result
    return result


def get_company_details(conn, company_id):
    if not company_id:
        return {}

    cursor = get_db_cursor(conn, dictionary=True)
    cursor.execute(
        """
        SELECT name, address, district, state, gstNo, gstType, gstSupplyType, pin
        FROM companies
        WHERE id = %s
        LIMIT 1
        """,
        (company_id,),
    )
    row = cursor.fetchone()
    cursor.close()
    return row or {}


def sanitize_party_address_line(value):
    cleaned = re.sub(r"</?ADDRESS\b[^>]*>", "", str(value or ""), flags=re.IGNORECASE)
    cleaned = cleaned.replace("&amp;", "&")
    cleaned = re.sub(r"[\x00-\x1F\x7F-\x9F]+", " ", cleaned)
    cleaned = re.sub(r"\]\s*\[|\[\s*\]", ", ", cleaned)
    cleaned = re.sub(r"[\[\]{}()]+", " ", cleaned)
    cleaned = re.sub(r"[|;:]+", ", ", cleaned)
    cleaned = re.sub(r"\s*[-–—]\s*", ", ", cleaned)
    cleaned = re.sub(r"\s*,\s*", ", ", cleaned)
    cleaned = re.sub(r",\s*,+", ", ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,")
    return cleaned


def build_party_address_lines(company_row):
    if not company_row:
        return []

    address_lines = []

    primary_address = str(company_row.get("address") or "").strip()
    if primary_address:
        split_lines = [sanitize_party_address_line(part) for part in re.split(r"[\r\n]+", primary_address) if sanitize_party_address_line(part)]
        address_lines.extend(split_lines)

    location_parts = [
        str(company_row.get("district") or "").strip(),
        str(company_row.get("state") or "").strip(),
        str(company_row.get("pin") or "").strip(),
        "India",
    ]
    location_line = ", ".join([part for part in location_parts if part])
    if location_line and location_line not in address_lines:
        address_lines.append(location_line)

    unique_lines = []
    for line in address_lines:
        cleaned_line = sanitize_party_address_line(line)
        if cleaned_line and cleaned_line not in unique_lines:
            unique_lines.append(cleaned_line)
    return unique_lines


def build_party_details(customer_name, company_row):
    tally_ledger_details, tally_error = fetch_tally_ledger_details(customer_name)
    if tally_error:
        log_terminal("PARTY", f"{customer_name}: {tally_error}. Falling back to company table values.")

    party_name = str((tally_ledger_details or {}).get("name") or customer_name).strip()
    party_mailing_name = str(
        (tally_ledger_details or {}).get("mailing_name")
        or (company_row or {}).get("name")
        or customer_name
    ).strip()
    party_state = str((tally_ledger_details or {}).get("state") or (company_row or {}).get("state") or "").strip()
    party_country = str((tally_ledger_details or {}).get("country") or "India").strip()
    party_pin = str((tally_ledger_details or {}).get("pin") or (company_row or {}).get("pin") or "").strip()
    party_gstin = str((tally_ledger_details or {}).get("gstin") or (company_row or {}).get("gstNo") or "").strip()
    party_gst_registration_type = str(
        (tally_ledger_details or {}).get("gst_registration_type")
        or resolve_party_gst_registration_type(company_row)
    ).strip()
    party_address_lines = (
        (tally_ledger_details or {}).get("address_lines")
        or build_party_address_lines(company_row)
    )

    return {
        "party_name": party_name,
        "mailing_name": party_mailing_name,
        "state": party_state,
        "country": party_country or "India",
        "pin": party_pin,
        "gstin": party_gstin,
        "gst_registration_type": party_gst_registration_type,
        "address_lines": party_address_lines,
    }


def resolve_party_gst_registration_type(company_row):
    gst_type = str((company_row or {}).get("gstType") or "").strip()
    gst_no = str((company_row or {}).get("gstNo") or "").strip()

    normalized_map = {
        "REGULAR": "Regular",
        "COMPOSITION": "Composition",
        "CONSUMER": "Consumer",
        "UNREGISTERED": "Unregistered",
        "SEZ": "SEZ",
        "SEZ UNIT": "SEZ",
        "SEZ DEVELOPER": "SEZ",
        "EXEMPT": "Consumer",
    }

    if gst_type:
        normalized = normalized_map.get(gst_type.upper())
        if normalized:
            return normalized
        return gst_type

    if gst_no:
        return "Regular"

    return "Unregistered"


def normalize_item_source(value):
    normalized = str(value or "").strip().upper()
    if normalized == "PHP":
        return "PHP"
    if normalized == "PLATE":
        return "PLATE"
    if normalized == "MATERIAL":
        return "MATERIAL"
    return "FG"


def resolve_effective_line_uom(line):
    item_source = normalize_item_source(line.get("itemSource"))
    raw_uom = normalize_uom(line.get("uom"))
    if raw_uom:
        return raw_uom
    if item_source in {"PHP", "PLATE"}:
        return "PCS"
    return ""


def resolve_invoice_line_item_details(cursor, item_source, item_id, npd_id):
    item_name = ""
    uom = ""
    npd_part = ""

    lookup_npd_id = str(npd_id or item_id or "").strip()
    lookup_item_id = str(item_id or "").strip()

    if item_source == "FG" and lookup_npd_id:
        cursor.execute(
            """
            SELECT itemName, uom, part
            FROM npd
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_npd_id,),
        )
        npd_row = cursor.fetchone()
        if npd_row:
            item_name = npd_row.get("itemName") or ""
            uom = npd_row.get("uom") or ""
            npd_part = npd_row.get("part") or ""

    elif item_source == "PHP" and lookup_item_id:
        cursor.execute(
            """
            SELECT itemName
            FROM php_item_master
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        php_row = cursor.fetchone()
        if php_row:
            item_name = php_row.get("itemName") or ""
            uom = "PCS"

    elif item_source == "PLATE" and lookup_item_id:
        cursor.execute(
            """
            SELECT itemName
            FROM plate_item_master
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        plate_row = cursor.fetchone()
        if plate_row:
            item_name = plate_row.get("itemName") or ""
            uom = "PCS"

    elif item_source == "MATERIAL" and lookup_item_id:
        cursor.execute(
            """
            SELECT name, uom
            FROM materials
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        material_row = cursor.fetchone()
        if material_row:
            item_name = material_row.get("name") or ""
            uom = material_row.get("uom") or ""

    if not item_name and lookup_npd_id:
        cursor.execute(
            """
            SELECT itemName, uom, part
            FROM npd
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_npd_id,),
        )
        npd_row = cursor.fetchone()
        if npd_row:
            item_name = npd_row.get("itemName") or ""
            if not uom:
                uom = npd_row.get("uom") or ""
            npd_part = npd_row.get("part") or ""

    if not item_name and lookup_item_id and item_source != "PHP":
        cursor.execute(
            """
            SELECT itemName
            FROM php_item_master
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        php_row = cursor.fetchone()
        if php_row:
            item_name = php_row.get("itemName") or ""
            if not uom:
                uom = "PCS"

    if not item_name and lookup_item_id and item_source != "PLATE":
        cursor.execute(
            """
            SELECT itemName
            FROM plate_item_master
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        plate_row = cursor.fetchone()
        if plate_row:
            item_name = plate_row.get("itemName") or ""
            if not uom:
                uom = "PCS"

    if not item_name and lookup_item_id and item_source != "MATERIAL":
        cursor.execute(
            """
            SELECT name, uom
            FROM materials
            WHERE id = %s
            LIMIT 1
            """,
            (lookup_item_id,),
        )
        material_row = cursor.fetchone()
        if material_row:
            item_name = material_row.get("name") or ""
            if not uom:
                uom = material_row.get("uom") or ""

    return item_name, uom, npd_part


def get_invoice_lines(conn, invoice_id):
    cursor = get_db_cursor(conn, dictionary=True)
    cursor.execute(
        """
        SELECT *
        FROM invoice_line_items
        WHERE invoiceId = %s
        ORDER BY id
        """,
        (invoice_id,),
    )
    rows = cursor.fetchall()

    processed_lines = []
    for row in rows:
        item_id = row.get("itemId")
        npd_id = row.get("npdId")
        item_source = normalize_item_source(row.get("itemSource"))
        item_name, uom, npd_part = resolve_invoice_line_item_details(cursor, item_source, item_id, npd_id)

        processed_lines.append(
            {
                "id": row.get("id"),
                "invoiceId": row.get("invoiceId"),
                "loadingSlipId": row.get("loadingSlipId"),
                "itemId": item_id,
                "itemSource": item_source,
                "npdId": npd_id,
                "itemName": item_name or "Unknown Item",
                "uom": normalize_uom(uom) or ("PCS" if item_source in {"PHP", "PLATE"} else ""),
                "npdPartNo": str(npd_part or "").strip(),
                "qty": to_float(row.get("qty")),
                "rate": to_float(row.get("rate")),
                "amount": to_float(row.get("amount")),
                "gstRate": to_float(row.get("gstRate")),
                "cgst": to_float(row.get("cgst")),
                "sgst": to_float(row.get("sgst")),
                "igst": to_float(row.get("igst")),
            }
        )

    cursor.close()
    return processed_lines


def get_invoice_dispatch_details(conn, invoice_id, item_lines):
    loading_slip_ids = sorted(
        {
            str(line.get("loadingSlipId") or "").strip()
            for line in item_lines
            if str(line.get("loadingSlipId") or "").strip()
        }
    )

    slip_nos = []
    truck_nos = []
    po_numbers = []
    order_dates = []
    order_details = []
    transporter = ""
    dispatch_plan_ids = []
    cursor = get_db_cursor(conn, dictionary=True)

    try:
        if loading_slip_ids:
            placeholders = ", ".join(["%s"] * len(loading_slip_ids))
            cursor.execute(
                f"""
                SELECT ls.id, ls.slipNo, ls.lines, tr.truckNo
                FROM loading_slips ls
                LEFT JOIN trucks tr ON tr.id = ls.truckId
                WHERE ls.id IN ({placeholders})
                """,
                tuple(loading_slip_ids),
            )
            for row in cursor.fetchall():
                slip_no = str(row.get("slipNo") or "").strip()
                truck_no = str(row.get("truckNo") or "").strip()
                raw_lines = row.get("lines")
                if slip_no and slip_no not in slip_nos:
                    slip_nos.append(slip_no)
                if truck_no and truck_no not in truck_nos:
                    truck_nos.append(truck_no)
                if raw_lines:
                    try:
                        slip_lines = raw_lines if isinstance(raw_lines, list) else json.loads(raw_lines)
                    except Exception:
                        slip_lines = []
                    for slip_line in slip_lines or []:
                        dispatch_plan_id = str((slip_line or {}).get("dispatchPlanId") or "").strip()
                        if dispatch_plan_id and dispatch_plan_id not in dispatch_plan_ids:
                            dispatch_plan_ids.append(dispatch_plan_id)

        if not truck_nos or not transporter:
            gate_pass_fields = ["truckNo"]
            if column_exists(conn, "gate_passes", "transporter"):
                gate_pass_fields.append("transporter")
            gate_pass_select = ", ".join(gate_pass_fields)
            cursor.execute(
                f"""
                SELECT {gate_pass_select}
                FROM gate_passes
                WHERE invoiceId = %s
                """,
                (invoice_id,),
            )
            for row in cursor.fetchall():
                truck_no = str(row.get("truckNo") or "").strip()
                transporter_name = str(row.get("transporter") or "").strip()
                if truck_no and truck_no not in truck_nos:
                    truck_nos.append(truck_no)
                if transporter_name and not transporter:
                    transporter = transporter_name

        if dispatch_plan_ids:
            placeholders = ", ".join(["%s"] * len(dispatch_plan_ids))
            cursor.execute(
                f"""
                SELECT DISTINCT o.poNumber, o.orderDate
                FROM dispatch_plans dp
                INNER JOIN orders_schedule os ON os.id = dp.scheduleId
                INNER JOIN orders o ON o.id = os.orderId
                WHERE dp.id IN ({placeholders})
                ORDER BY o.orderDate, o.poNumber
                """,
                tuple(dispatch_plan_ids),
            )
            for row in cursor.fetchall():
                po_number = str(row.get("poNumber") or "").strip()
                order_date = str(row.get("orderDate") or "").strip()
                if po_number and po_number not in po_numbers:
                    po_numbers.append(po_number)
                if order_date and order_date not in order_dates:
                    order_dates.append(order_date)
                if po_number:
                    order_entry = {
                        "poNumber": po_number,
                        "orderDate": order_date,
                    }
                    if order_entry not in order_details:
                        order_details.append(order_entry)
    finally:
        cursor.close()

    return {
        "loadingSlipNos": slip_nos,
        "truckNos": truck_nos,
        "orderNos": po_numbers,
        "orderDates": order_dates,
        "orderDetails": order_details,
        "transporter": transporter,
    }


def build_invoice_narration(dispatch_details):
    parts = []
    slip_nos = dispatch_details.get("loadingSlipNos") or []
    truck_nos = dispatch_details.get("truckNos") or []

    if slip_nos:
        parts.append(f"Loading Slips: {', '.join(slip_nos)}")
    if truck_nos:
        parts.append(f"Truck No: {', '.join(truck_nos)}")

    return " | ".join(parts)


def validate_invoice_lines(item_lines):
    errors = []

    for index, line in enumerate(item_lines, start=1):
        item_name = line.get("itemName") or f"Line {index}"
        effective_uom = resolve_effective_line_uom(line)

        if to_float(line.get("qty")) <= 0:
            errors.append(f"{item_name}: qty missing or zero")
        if to_float(line.get("rate")) <= 0:
            errors.append(f"{item_name}: rate missing or zero")
        if to_float(line.get("amount")) <= 0:
            errors.append(f"{item_name}: amount missing or zero")
        if not effective_uom:
            errors.append(f"{item_name}: uom missing")
        if item_name == "Unknown Item":
            errors.append(f"Line {index}: item name not found")

    return errors


def resolve_sales_ledger_name(invoice_row, company_row, item_lines):
    gst_supply_type = str((company_row or {}).get("gstSupplyType") or "").strip().upper()
    igst_amount = round(to_float(invoice_row.get("igst")), 2)

    candidate_rates = []
    header_gst_rate = round(to_float(invoice_row.get("gstRate")), 2)
    if header_gst_rate > 0:
        candidate_rates.append(header_gst_rate)

    for line in item_lines:
        line_gst_rate = round(to_float(line.get("gstRate")), 2)
        if line_gst_rate > 0:
            candidate_rates.append(line_gst_rate)

    effective_gst_rate = max(candidate_rates) if candidate_rates else 0.0

    if abs(effective_gst_rate - 5.0) < 0.01:
        return SALES_5_LEDGER_NAME

    if abs(effective_gst_rate - 18.0) < 0.01:
        return SALES_18_LEDGER_NAME

    return SALES_LEDGER_NAME


def format_tax_rate(rate):
    value = to_float(rate)
    if value.is_integer():
        return f"{value:.1f}%"
    return f"{value:g}%"


def resolve_tax_ledger_name(prefix, rate, fallback_name):
    rate_value = round(to_float(rate), 2)
    if rate_value <= 0:
        return fallback_name
    return f"{prefix} {format_tax_rate(rate_value)}"


def derive_tax_rates(invoice_row, item_lines):
    header_gst_rate = round(to_float(invoice_row.get("gstRate")), 2)
    if header_gst_rate > 0:
        return {
            "cgst_rate": header_gst_rate / 2,
            "sgst_rate": header_gst_rate / 2,
            "igst_rate": header_gst_rate,
        }

    line_gst_rates = sorted(
        {
            round(to_float(line.get("gstRate")), 2)
            for line in item_lines
            if round(to_float(line.get("gstRate")), 2) > 0
        }
    )

    if line_gst_rates:
        primary_rate = line_gst_rates[0]
        return {
            "cgst_rate": primary_rate / 2,
            "sgst_rate": primary_rate / 2,
            "igst_rate": primary_rate,
        }

    other_charges_gst_rate = round(to_float(invoice_row.get("otherChargesGstRate")), 2)
    if other_charges_gst_rate > 0:
        return {
            "cgst_rate": other_charges_gst_rate / 2,
            "sgst_rate": other_charges_gst_rate / 2,
            "igst_rate": other_charges_gst_rate,
        }

    return {
        "cgst_rate": 0,
        "sgst_rate": 0,
        "igst_rate": 0,
    }


def compute_effective_round_off(invoice_row, total_item_amount=0.0):
    return round(to_float(invoice_row.get("roundOff")), 2)


def create_sales_voucher_xml(
    invoice_row,
    customer_name,
    company_row,
    item_lines,
    sales_ledger_name,
    narration_text="",
    dispatch_details=None,
):
    invoice_date = format_tally_date(invoice_row.get("date"))
    invoice_no = str(invoice_row.get("invoiceNo") or "").strip()
    customer_name = customer_name or "Unknown Customer"
    dispatch_details = dispatch_details or {}

    cgst = round(to_float(invoice_row.get("cgst")), 2)
    sgst = round(to_float(invoice_row.get("sgst")), 2)
    igst = round(to_float(invoice_row.get("igst")), 2)
    other_charges = round(to_float(invoice_row.get("otherCharges")), 2)
    derived_tax_rates = derive_tax_rates(invoice_row, item_lines)
    cgst_rate = derived_tax_rates["cgst_rate"] if cgst > 0 else 0
    sgst_rate = derived_tax_rates["sgst_rate"] if sgst > 0 else 0
    igst_rate = derived_tax_rates["igst_rate"] if igst > 0 else 0
    cgst_ledger_name = resolve_tax_ledger_name(CGST_LEDGER_PREFIX, cgst_rate, CGST_LEDGER_NAME)
    sgst_ledger_name = resolve_tax_ledger_name(SGST_LEDGER_PREFIX, sgst_rate, SGST_LEDGER_NAME)
    igst_ledger_name = resolve_tax_ledger_name(IGST_LEDGER_PREFIX, igst_rate, IGST_LEDGER_NAME)

    inventory_xml = ""
    total_item_amount = 0.0

    for line in item_lines:
        item_name = line.get("itemName") or ""
        qty = to_float(line.get("qty"))
        rate = to_float(line.get("rate"))
        amount = round(to_float(line.get("amount")) or qty * rate, 2)
        uom = resolve_effective_line_uom(line)
        qty_text = f"{qty:g} {esc(uom)}".strip()
        rate_text = f"{rate:g}/{esc(uom)}" if uom else f"{rate:g}"

        total_item_amount += amount

        inventory_xml += f"""
                        <ALLINVENTORYENTRIES.LIST>
                            <STOCKITEMNAME>{esc(item_name)}</STOCKITEMNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
                            <ACTUALQTY>{qty_text}</ACTUALQTY>
                            <BILLEDQTY>{qty_text}</BILLEDQTY>
                            <RATE>{rate_text}</RATE>
                            <AMOUNT>{amount:.2f}</AMOUNT>
                            <ACCOUNTINGALLOCATIONS.LIST>
                                <LEDGERNAME>{esc(sales_ledger_name)}</LEDGERNAME>
                                <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                                <AMOUNT>{amount:.2f}</AMOUNT>
                            </ACCOUNTINGALLOCATIONS.LIST>
                        </ALLINVENTORYENTRIES.LIST>
"""

    total_item_amount = round(total_item_amount, 2)
    round_off = compute_effective_round_off(invoice_row, total_item_amount)
    total_invoice_amount = round(total_item_amount + cgst + sgst + igst + other_charges + round_off, 2)

    ledger_entries_xml = f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(customer_name)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
                            <AMOUNT>-{total_invoice_amount:.2f}</AMOUNT>
                        </LEDGERENTRIES.LIST>
    """

    if cgst > 0:
        ledger_entries_xml += f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(cgst_ledger_name)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <AMOUNT>{cgst:.2f}</AMOUNT>
                        </LEDGERENTRIES.LIST>
        """

    if sgst > 0:
        ledger_entries_xml += f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(sgst_ledger_name)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <AMOUNT>{sgst:.2f}</AMOUNT>
                        </LEDGERENTRIES.LIST>
        """

    if igst > 0:
        ledger_entries_xml += f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(igst_ledger_name)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <AMOUNT>{igst:.2f}</AMOUNT>
                        </LEDGERENTRIES.LIST>
        """

    if other_charges != 0:
        ledger_entries_xml += f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(OTHER_CHARGES_LEDGER_NAME)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <AMOUNT>{other_charges:.2f}</AMOUNT>
                        </LEDGERENTRIES.LIST>
        """

    if round_off != 0:
        round_off_amount = f"{round_off:.2f}"
        ledger_entries_xml += f"""
                        <LEDGERENTRIES.LIST>
                            <LEDGERNAME>{esc(ROUND_OFF_LEDGER_NAME)}</LEDGERNAME>
                            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                            <AMOUNT>{round_off_amount}</AMOUNT>
                        </LEDGERENTRIES.LIST>
        """

    dispatch_doc_no = join_unique_values(dispatch_details.get("loadingSlipNos"))
    vehicle_no = join_unique_values(dispatch_details.get("truckNos"))
    order_nos = join_unique_values(dispatch_details.get("orderNos"))
    order_dates = dispatch_details.get("orderDates") or []
    order_date = format_tally_date(order_dates[0]) if order_dates else ""
    order_details = dispatch_details.get("orderDetails") or []
    destination = str(invoice_row.get("destination") or dispatch_details.get("destination") or "").strip()
    transporter = str(invoice_row.get("transporter") or dispatch_details.get("transporter") or "").strip()
    dispatch_through = "By Road"
    party_details = build_party_details(customer_name, company_row)
    party_name = str(party_details.get("party_name") or customer_name).strip()
    party_mailing_name = str(party_details.get("mailing_name") or customer_name).strip()
    party_state = str(party_details.get("state") or "").strip()
    party_country = str(party_details.get("country") or "India").strip()
    party_pin = str(party_details.get("pin") or "").strip()
    party_gstin = str(party_details.get("gstin") or "").strip()
    party_gst_registration_type = str(party_details.get("gst_registration_type") or "").strip()
    party_address_lines = party_details.get("address_lines") or []
    party_address_xml = (
        "\n".join(f"                            <ADDRESS>{esc(line)}</ADDRESS>" for line in party_address_lines)
        if party_address_lines
        else ""
    )
    basic_buyer_address_xml = (
        "\n".join(f"                            <BASICBUYERADDRESS>{esc(line)}</BASICBUYERADDRESS>" for line in party_address_lines)
        if party_address_lines
        else ""
    )
    consignee_address_xml = (
        "\n".join(f"                            <ADDRESS>{esc(line)}</ADDRESS>" for line in party_address_lines)
        if party_address_lines
        else ""
    )

    if order_details:
        invoice_order_list_xml = "".join(
            f"""
                        <INVOICEORDERLIST.LIST>
                            <BASICORDERDATE>{esc(format_tally_date(order_detail.get("orderDate") or ""))}</BASICORDERDATE>
                            <BASICPURCHASEORDERNO>{esc(order_detail.get("poNumber") or "")}</BASICPURCHASEORDERNO>
                        </INVOICEORDERLIST.LIST>"""
            for order_detail in order_details
            if str(order_detail.get("poNumber") or "").strip()
        )
    elif order_nos:
        invoice_order_list_xml = f"""
                        <INVOICEORDERLIST.LIST>
                            <BASICORDERDATE>{esc(order_date)}</BASICORDERDATE>
                            <BASICPURCHASEORDERNO>{esc(order_nos)}</BASICPURCHASEORDERNO>
                        </INVOICEORDERLIST.LIST>"""
    else:
        invoice_order_list_xml = ""

    return f"""
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Import Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <IMPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Vouchers</REPORTNAME>
                <STATICVARIABLES>
                    <SVCURRENTCOMPANY>{esc(TALLY_COMPANY_NAME)}</SVCURRENTCOMPANY>
                </STATICVARIABLES>
            </REQUESTDESC>
            <REQUESTDATA>
                <TALLYMESSAGE xmlns:UDF="TallyUDF">
                    <VOUCHER VCHTYPE="{esc(VOUCHER_TYPE_NAME)}" ACTION="Create">
                        <DATE>{invoice_date}</DATE>
                        <VOUCHERNUMBER>{esc(invoice_no)}</VOUCHERNUMBER>
                        <VOUCHERTYPENAME>{esc(VOUCHER_TYPE_NAME)}</VOUCHERTYPENAME>
                        <PARTYNAME>{esc(party_name)}</PARTYNAME>
                        <PARTYLEDGERNAME>{esc(customer_name)}</PARTYLEDGERNAME>
                        <BASICBUYERNAME>{esc(party_name)}</BASICBUYERNAME>
                        <PARTYMAILINGNAME>{esc(party_mailing_name)}</PARTYMAILINGNAME>
                        <PARTYGSTIN>{esc(party_gstin)}</PARTYGSTIN>
                        <GSTREGISTRATIONTYPE>{esc(party_gst_registration_type)}</GSTREGISTRATIONTYPE>
                        <GSTBUYERNAME>{esc(party_name)}</GSTBUYERNAME>
                        <GSTBUYERMAILINGNAME>{esc(party_mailing_name)}</GSTBUYERMAILINGNAME>
                        <GSTBUYERSTATE>{esc(party_state)}</GSTBUYERSTATE>
                        <GSTBUYERPINCODE>{esc(party_pin)}</GSTBUYERPINCODE>
                        <GSTBUYERGSTIN>{esc(party_gstin)}</GSTBUYERGSTIN>
                        <STATENAME>{esc(party_state)}</STATENAME>
                        <COUNTRYOFRESIDENCE>{esc(party_country)}</COUNTRYOFRESIDENCE>
                        <CONSIGNEENAME>{esc(party_name)}</CONSIGNEENAME>
                        <CONSIGNEEMAILINGNAME>{esc(party_mailing_name)}</CONSIGNEEMAILINGNAME>
                        <CONSIGNEESTATENAME>{esc(party_state)}</CONSIGNEESTATENAME>
                        <CONSIGNEECOUNTRYNAME>{esc(party_country)}</CONSIGNEECOUNTRYNAME>
                        <CONSIGNEEGSTIN>{esc(party_gstin)}</CONSIGNEEGSTIN>
                        <CONSIGNEEPINCODE>{esc(party_pin)}</CONSIGNEEPINCODE>
                        <CONSIGNEEPINNUMBER>{esc(party_pin)}</CONSIGNEEPINNUMBER>
                        <GSTCONSIGNEENAME>{esc(party_name)}</GSTCONSIGNEENAME>
                        <GSTCONSIGNEEMAILINGNAME>{esc(party_mailing_name)}</GSTCONSIGNEEMAILINGNAME>
                        <GSTCONSIGNEESTATE>{esc(party_state)}</GSTCONSIGNEESTATE>
                        <GSTCONSIGNEEPINCODE>{esc(party_pin)}</GSTCONSIGNEEPINCODE>
                        <GSTCONSIGNEEGSTIN>{esc(party_gstin)}</GSTCONSIGNEEGSTIN>
                        <PLACEOFSUPPLY>{esc(party_state)}</PLACEOFSUPPLY>
                        <PARTYPINCODE>{esc(party_pin)}</PARTYPINCODE>
                        <PARTYPINNUMBER>{esc(party_pin)}</PARTYPINNUMBER>
                        <BUYERPINNUMBER>{esc(party_pin)}</BUYERPINNUMBER>
                        <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>
                        <ISINVOICE>Yes</ISINVOICE>
                        <NARRATION>{esc(narration_text)}</NARRATION>
                        <ADDRESS.LIST TYPE="String">
{party_address_xml}
                        </ADDRESS.LIST>
                        <BASICBUYERADDRESS.LIST TYPE="String">
{basic_buyer_address_xml}
                        </BASICBUYERADDRESS.LIST>
                        <GSTBUYERADDRESS.LIST TYPE="String">
{party_address_xml}
                        </GSTBUYERADDRESS.LIST>
                        <CONSIGNEEADDRESS.LIST TYPE="String">
{consignee_address_xml}
                        </CONSIGNEEADDRESS.LIST>
                        <GSTCONSIGNEEADDRESS.LIST TYPE="String">
{consignee_address_xml}
                        </GSTCONSIGNEEADDRESS.LIST>
                        <BASICSHIPTOADDRESS.LIST TYPE="String">
{consignee_address_xml}
                        </BASICSHIPTOADDRESS.LIST>
                        <BASICCONSIGNEEADDRESS.LIST TYPE="String">
{consignee_address_xml}
                        </BASICCONSIGNEEADDRESS.LIST>
                        <TRANSPORTERNAME>{esc(dispatch_through)}</TRANSPORTERNAME>
                        <TRANSPORTMODE>Road</TRANSPORTMODE>
                        <DESPATCHEDTHROUGH>{esc(dispatch_through)}</DESPATCHEDTHROUGH>
                        <CARRIERNAME>{esc(transporter)}</CARRIERNAME>
                        <CARRIERNAMEAGENCY>{esc(transporter)}</CARRIERNAMEAGENCY>
                        <EICHECKPOST>{esc(transporter)}</EICHECKPOST>
                        <BASICSHIPPEDBY>{esc(dispatch_through)}</BASICSHIPPEDBY>
                        <BASICSHIPDOCUMENTNO>{esc(dispatch_doc_no)}</BASICSHIPDOCUMENTNO>
                        <SHIPPEDVIA>{esc(transporter or dispatch_through)}</SHIPPEDVIA>
                        <BASICSHIPMETHOD>{esc(dispatch_through)}</BASICSHIPMETHOD>
                        <BASICFINALDESTINATION>{esc(destination)}</BASICFINALDESTINATION>
                        <BASICSHIPFORWARDER>{esc(transporter)}</BASICSHIPFORWARDER>
                        <BASICSHIPVESSELNO>{esc(vehicle_no)}</BASICSHIPVESSELNO>
                        <BASICORDERREF>{esc(order_nos)}</BASICORDERREF>
                        <BASICORDERDATE>{esc(order_date)}</BASICORDERDATE>
                        {invoice_order_list_xml}
                        <INVOICESHIPLIST.LIST>
                            <BASICSHIPMETHOD>{esc(dispatch_through)}</BASICSHIPMETHOD>
                            <BASICSHIPFORWARDER>{esc(transporter)}</BASICSHIPFORWARDER>
                        </INVOICESHIPLIST.LIST>
                        <BASICSHIPDELIVERYPROPERTIES.LIST>
                            <BASICSHIPMETHOD>{esc(dispatch_through)}</BASICSHIPMETHOD>
                            <BASICSHIPFORWARDER>{esc(transporter)}</BASICSHIPFORWARDER>
                        </BASICSHIPDELIVERYPROPERTIES.LIST>
                        <EWAYBILLDETAILS.LIST>
                            <TRANSPORTERNAME>{esc(transporter)}</TRANSPORTERNAME>
                            <TRANSMODE>Road</TRANSMODE>
                        </EWAYBILLDETAILS.LIST>
                        {inventory_xml}
                        {ledger_entries_xml}
                    </VOUCHER>
                </TALLYMESSAGE>
            </REQUESTDATA>
        </IMPORTDATA>
    </BODY>
</ENVELOPE>
"""


def post_to_tally(xml_data):
    response_text = tally_request(xml_data)

    if response_text and "Could not set &apos;SVCurrentCompany&apos;" in response_text:
        fallback_xml = re.sub(
            r"\s*<STATICVARIABLES>\s*<SVCURRENTCOMPANY>.*?</SVCURRENTCOMPANY>\s*</STATICVARIABLES>",
            "",
            xml_data,
            flags=re.IGNORECASE | re.DOTALL,
        )
        log_terminal("RETRY", "Retrying Tally post without SVCURRENTCOMPANY")
        return tally_request(fallback_xml)

    return response_text


def is_tally_success(response_text):
    if not response_text:
        return False
    return "<CREATED>1</CREATED>" in response_text or "<ALTERED>1</ALTERED>" in response_text


def extract_tally_error(response_text):
    if not response_text:
        return "Empty response from Tally"

    match = re.search(r"<LINEERROR>(.*?)</LINEERROR>", response_text, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()

    if "<EXCEPTIONS>0</EXCEPTIONS>" not in response_text and "<EXCEPTIONS>" in response_text:
        return "Tally returned an exception. Check response for details."

    return response_text[:200]


def validate_tally_masters(customer_name, sales_ledger_name, item_lines, invoice_row):
    errors = []

    master_checks = [("Ledger", customer_name), ("Ledger", sales_ledger_name)]

    other_charges = round(to_float(invoice_row.get("otherCharges")), 2)
    cgst = round(to_float(invoice_row.get("cgst")), 2)
    sgst = round(to_float(invoice_row.get("sgst")), 2)
    igst = round(to_float(invoice_row.get("igst")), 2)
    total_item_amount = round(sum(round(to_float(line.get("amount")), 2) for line in item_lines), 2)
    round_off = compute_effective_round_off(invoice_row, total_item_amount)
    derived_tax_rates = derive_tax_rates(invoice_row, item_lines)

    if other_charges != 0:
        master_checks.append(("Ledger", OTHER_CHARGES_LEDGER_NAME))
    if round_off != 0:
        master_checks.append(("Ledger", ROUND_OFF_LEDGER_NAME))
    if cgst > 0:
        master_checks.append(
            ("Ledger", resolve_tax_ledger_name(CGST_LEDGER_PREFIX, derived_tax_rates["cgst_rate"], CGST_LEDGER_NAME))
        )
    if sgst > 0:
        master_checks.append(
            ("Ledger", resolve_tax_ledger_name(SGST_LEDGER_PREFIX, derived_tax_rates["sgst_rate"], SGST_LEDGER_NAME))
        )
    if igst > 0:
        master_checks.append(
            ("Ledger", resolve_tax_ledger_name(IGST_LEDGER_PREFIX, derived_tax_rates["igst_rate"], IGST_LEDGER_NAME))
        )

    for line in item_lines:
        master_checks.append(("Stock Item", line.get("itemName") or ""))

    checked = set()
    for object_type, object_name in master_checks:
        key = (object_type, object_name)
        if key in checked or not object_name:
            continue
        checked.add(key)
        exists, message = check_tally_object_exists(object_type, object_name)
        if not exists:
            errors.append(message)

    for line in item_lines:
        item_name = str(line.get("itemName") or "").strip()
        erp_uom = resolve_effective_line_uom(line)
        if not item_name or not erp_uom:
            continue

        tally_stock_item, error_message = fetch_tally_stock_item_details(item_name)
        if error_message:
            errors.append(error_message)
            continue
        tally_uom = tally_stock_item.get("base_uom") or ""
        if tally_uom != erp_uom:
            errors.append(f"{item_name}: UOM mismatch. ERP={erp_uom}, Tally={tally_uom}")

        if line.get("npdId"):
            erp_part_no = normalize_part_no(line.get("npdPartNo"))
            if erp_part_no:
                tally_part_no = normalize_part_no(tally_stock_item.get("part_no"))
                tally_aliases = {
                    normalize_part_no(value)
                    for value in (tally_stock_item.get("mailing_names") or [])
                    if normalize_part_no(value)
                }
                if not tally_part_no and erp_part_no not in tally_aliases:
                    errors.append(f"{item_name}: Part No missing in Tally. ERP={line.get('npdPartNo')}")
                elif erp_part_no != tally_part_no and erp_part_no not in tally_aliases:
                    display_part = tally_stock_item.get("part_no") or ", ".join(tally_stock_item.get("mailing_names") or [])
                    errors.append(
                        f"{item_name}: Part No mismatch. ERP={line.get('npdPartNo')}, Tally={display_part or 'blank'}"
                    )

    return errors


def prevalidate_pending_invoices(conn, pending_invoice_rows):
    valid_contexts = []
    precheck_summary = {
        "customers_ok": set(),
        "stock_items_ok": set(),
        "item_uom_ok": set(),
        "invoice_numbers_clear": set(),
        "invoice_numbers_existing": [],
        "npd_part_matches": set(),
    }
    halted_invoice_no = ""
    halted_remark = ""

    print("==========================================")
    print("Prechecking pending invoices before posting")
    print("==========================================")

    for invoice_row in pending_invoice_rows:
        invoice_id = invoice_row.get("id")
        invoice_no = str(invoice_row.get("invoiceNo") or "").strip()
        company_name = ""

        print(f"Precheck Invoice ID: {invoice_id} | Invoice No: {invoice_no}")

        try:
            if conn is None and use_tally_sync_api():
                invoice_context = get_invoice_context_api(invoice_id)
                invoice_row = invoice_context.get("invoiceRow") or invoice_row
                company_row = invoice_context.get("companyRow") or {}
                item_lines = invoice_context.get("itemLines") or []
                dispatch_details = invoice_context.get("dispatchDetails") or {}
            else:
                company_row = get_company_details(conn, invoice_row.get("companyId"))
                item_lines = get_invoice_lines(conn, invoice_id)
                dispatch_details = None

            company_name = (company_row or {}).get("name") or ""
            if not company_name:
                remark = "Company ledger not found in companies table"
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    remark,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                log_terminal("PRECHECK", remark)
                halted_invoice_no = invoice_no
                halted_remark = remark
                break

            if not item_lines:
                remark = "No invoice line items found"
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    remark,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                log_terminal("PRECHECK", remark)
                halted_invoice_no = invoice_no
                halted_remark = remark
                break

            line_errors = validate_invoice_lines(item_lines)
            if line_errors:
                remark = " | ".join(line_errors[:10])
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    remark,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                log_terminal("PRECHECK", remark)
                halted_invoice_no = invoice_no
                halted_remark = remark
                break

            if not invoice_no:
                remark = "Invoice No missing"
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    remark,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                log_terminal("PRECHECK", remark)
                halted_invoice_no = invoice_no
                halted_remark = remark
                break

            precheck_summary["customers_ok"].add(company_name)

            sales_ledger_name = resolve_sales_ledger_name(invoice_row, company_row, item_lines)
            if dispatch_details is None:
                dispatch_details = get_invoice_dispatch_details(conn, invoice_id, item_lines)
            narration_text = build_invoice_narration(dispatch_details)

            log_terminal("PRECHECK", f"{invoice_no}: checking saved Tally ID")
            existing_tally_id = str(invoice_row.get("tallyInvId") or "").strip()
            tally_reference = fetch_tally_voucher_by_id(existing_tally_id) if existing_tally_id else {}
            if existing_tally_id and tally_reference:
                remark = "Voucher already exists in Tally. Matched by saved Tally ID."
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    True,
                    remark,
                    invoice_row.get("updatedBy") or DEFAULT_UPDATED_BY,
                    tally_reference.get("tallyInvNo") or invoice_no,
                    format_iso_date(tally_reference.get("tallyInvDate")),
                    existing_tally_id,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                precheck_summary["invoice_numbers_existing"].append(invoice_no)
                log_terminal("PRECHECK", f"{invoice_no}: {remark}")
                continue

            log_terminal("PRECHECK", f"{invoice_no}: checking existing invoice number in Tally")
            voucher_by_number = fetch_tally_voucher_reference(
                invoice_no,
                VOUCHER_TYPE_NAME,
                invoice_row.get("date"),
                company_name,
            )
            if voucher_by_number:
                remark = "This Invoice already exists in tally."
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    True,
                    remark,
                    invoice_row.get("updatedBy") or DEFAULT_UPDATED_BY,
                    voucher_by_number.get("tallyInvNo") or invoice_no,
                    format_iso_date(voucher_by_number.get("tallyInvDate")),
                    voucher_by_number.get("tallyInvId"),
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                precheck_summary["invoice_numbers_existing"].append(invoice_no)
                log_terminal("PRECHECK", f"{invoice_no}: {remark}")
                continue

            log_terminal("PRECHECK", f"{invoice_no}: validating ledgers and stock items in Tally")
            tally_master_errors = validate_tally_masters(
                company_name,
                sales_ledger_name,
                item_lines,
                invoice_row,
            )
            if tally_master_errors:
                remark = " | ".join(tally_master_errors[:10])
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    remark,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="PRECHECK",
                )
                log_terminal("PRECHECK", remark)
                halted_invoice_no = invoice_no
                halted_remark = remark
                break

            for line in item_lines:
                item_name = str(line.get("itemName") or "").strip()
                item_uom = normalize_uom(line.get("uom"))
                if item_name:
                    precheck_summary["stock_items_ok"].add(item_name)
                if item_name and item_uom:
                    precheck_summary["item_uom_ok"].add((item_name, item_uom))
                if line.get("npdId"):
                    npd_part_no = normalize_part_no(line.get("npdPartNo"))
                    if item_name and npd_part_no:
                        precheck_summary["npd_part_matches"].add((item_name, npd_part_no))

            precheck_summary["invoice_numbers_clear"].add(invoice_no)
            valid_contexts.append(
                {
                    "invoice_row": invoice_row,
                    "invoice_id": invoice_id,
                    "invoice_no": invoice_no,
                    "company_row": company_row,
                    "company_name": company_name,
                    "item_lines": item_lines,
                    "sales_ledger_name": sales_ledger_name,
                    "dispatch_details": dispatch_details,
                    "narration_text": narration_text,
                }
            )
        except Exception as exc:
            error_message = str(exc)[:1000]
            update_invoice_tally_status(
                conn,
                invoice_id,
                False,
                error_message,
                invoice_row=invoice_row,
                company_name=company_name,
                stage="PRECHECK",
            )
            log_terminal("PRECHECK", error_message)
            halted_invoice_no = invoice_no
            halted_remark = error_message
            break

    print("==========================================")
    print("Precheck Summary Before Posting")
    print("==========================================")
    if precheck_summary["customers_ok"]:
        print(f"[OK] All Customers exist in Tally for {len(precheck_summary['customers_ok'])} customer ledger(s).")
    else:
        print("[INFO] No customer ledgers were cleared in precheck.")

    if precheck_summary["stock_items_ok"]:
        print(f"[OK] All Stock Items exist in Tally for {len(precheck_summary['stock_items_ok'])} unique item name(s).")
    else:
        print("[INFO] No stock items were cleared in precheck.")

    if precheck_summary["item_uom_ok"]:
        print(f"[OK] All Items exist with matching Unit for {len(precheck_summary['item_uom_ok'])} unique item/unit combination(s).")
    else:
        print("[INFO] No item/unit combinations were cleared in precheck.")

    if precheck_summary["invoice_numbers_clear"]:
        print(f"[OK] There is no existing Invoice No. in Tally for {len(precheck_summary['invoice_numbers_clear'])} invoice(s) cleared for posting.")
    else:
        print("[INFO] No invoice numbers were cleared for posting.")

    if precheck_summary["invoice_numbers_existing"]:
        print(f"[INFO] Existing Invoice No. already found in Tally for {len(precheck_summary['invoice_numbers_existing'])} invoice(s).")

    if precheck_summary["npd_part_matches"]:
        print(
            f"[OK] NPD Stock Items exist with matching Name and Part No for "
            f"{len(precheck_summary['npd_part_matches'])} unique item/part combination(s)."
        )
    else:
        print("[INFO] No NPD Part No comparisons were applicable in this precheck.")

    if halted_invoice_no:
        print(
            f"[STOP] Precheck halted at Invoice No. {halted_invoice_no}. "
            f"No later invoices will be posted in this run. Reason: {halted_remark}"
        )

    return valid_contexts


def update_invoice_tally_status(
    conn,
    invoice_id,
    success,
    remark,
    tally_by=None,
    tally_inv_no=None,
    tally_inv_date=None,
    tally_inv_id=None,
    invoice_row=None,
    company_name="",
    stage="SYNC",
):
    if conn is None and use_tally_sync_api():
        update_invoice_tally_status_api(
            invoice_id,
            success,
            remark,
            tally_by=tally_by,
            tally_inv_no=tally_inv_no,
            tally_inv_date=tally_inv_date,
            tally_inv_id=tally_inv_id,
        )
        append_excel_log(
            "SUCCESS" if success else "FAILURE",
            stage,
            remark,
            invoice_row=invoice_row or {"id": invoice_id},
            company_name=company_name,
            tally_inv_no=tally_inv_no,
            tally_inv_date=tally_inv_date,
            tally_inv_id=tally_inv_id,
        )
        return

    cursor = get_db_cursor(conn)
    if success:
        sql = """
            UPDATE invoices
            SET
                tallyTimestamp = %s,
                tallyBy = %s,
                tallySyncRemark = %s,
                tallyInvNo = %s,
                tallyInvDate = %s,
                tallyInvId = %s
            WHERE id = %s
            """
        cursor.execute(
            sql,
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                tally_by or DEFAULT_UPDATED_BY,
                remark,
                tally_inv_no,
                tally_inv_date,
                tally_inv_id,
                invoice_id,
            ),
        )
    else:
        cursor.execute(
            """
            UPDATE invoices
            SET tallySyncRemark = %s
            WHERE id = %s
            """,
            (remark, invoice_id),
        )
    conn.commit()
    cursor.close()
    append_excel_log(
        "SUCCESS" if success else "FAILURE",
        stage,
        remark,
        invoice_row=invoice_row or {"id": invoice_id},
        company_name=company_name,
        tally_inv_no=tally_inv_no,
        tally_inv_date=tally_inv_date,
        tally_inv_id=tally_inv_id,
    )


def sync_invoices_to_tally():
    conn = None

    if use_tally_sync_api():
        pending_invoice_rows = get_pending_invoice_rows_api()
    else:
        conn = get_db_connection()
        ensure_invoice_sync_columns(conn)
        pending_invoice_rows = get_pending_invoice_rows(conn)

    try:
        print("==========================================")
        print(f"Pending invoices found: {len(pending_invoice_rows)}")
        print("==========================================")

        valid_contexts = prevalidate_pending_invoices(conn, pending_invoice_rows)

        print("==========================================")
        print(f"Invoices cleared for posting: {len(valid_contexts)}")
        print("==========================================")

        for context in valid_contexts:
            invoice_row = context["invoice_row"]
            invoice_id = context["invoice_id"]
            invoice_no = context["invoice_no"]

            print(f"\nProcessing Invoice ID: {invoice_id} | Invoice No: {invoice_no}")

            try:
                company_name = context["company_name"]
                company_row = context["company_row"]
                item_lines = context["item_lines"]
                sales_ledger_name = context["sales_ledger_name"]
                dispatch_details = context["dispatch_details"]
                narration_text = context["narration_text"]

                existing_tally_id = str(invoice_row.get("tallyInvId") or "").strip()
                tally_reference = fetch_tally_voucher_by_id(existing_tally_id) if existing_tally_id else {}
                if existing_tally_id and tally_reference:
                    remark = "Voucher already exists in Tally. Matched by saved Tally ID."
                    update_invoice_tally_status(
                        conn,
                        invoice_id,
                        True,
                        remark,
                        invoice_row.get("updatedBy") or DEFAULT_UPDATED_BY,
                        tally_reference.get("tallyInvNo") or invoice_no,
                        format_iso_date(tally_reference.get("tallyInvDate")),
                        existing_tally_id,
                        invoice_row=invoice_row,
                        company_name=company_name,
                        stage="SYNC",
                    )
                    print(f"Skipping creation: {remark} | {tally_reference}")
                    continue

                voucher_by_number = fetch_tally_voucher_reference(
                    invoice_no,
                    VOUCHER_TYPE_NAME,
                    invoice_row.get("date"),
                    company_name,
                )
                if voucher_by_number:
                    remark = "This Invoice already exists in tally."
                    update_invoice_tally_status(
                        conn,
                        invoice_id,
                        True,
                        remark,
                        invoice_row.get("updatedBy") or DEFAULT_UPDATED_BY,
                        voucher_by_number.get("tallyInvNo") or invoice_no,
                        format_iso_date(voucher_by_number.get("tallyInvDate")),
                        voucher_by_number.get("tallyInvId"),
                        invoice_row=invoice_row,
                        company_name=company_name,
                        stage="SYNC",
                    )
                    print(f"Skipping creation: {remark} | {voucher_by_number}")
                    continue

                for line in item_lines:
                    print(
                        "Line -> "
                        f"item={line.get('itemName')}, "
                        f"qty={line.get('qty')}, "
                        f"uom={line.get('uom')}, "
                        f"rate={line.get('rate')}, "
                        f"amount={line.get('amount')}"
                    )
                print(f"Using sales ledger: {sales_ledger_name}")
                if narration_text:
                    print(f"Narration: {narration_text}")

                tally_xml = create_sales_voucher_xml(
                    invoice_row,
                    company_name,
                    company_row,
                    item_lines,
                    sales_ledger_name,
                    narration_text,
                    dispatch_details,
                )

                if DEBUG_TALLY_XML:
                    print("Generated Tally XML:")
                    print(tally_xml)

                tally_response = post_to_tally(tally_xml)

                print("Tally Response:")
                print(tally_response)

                if is_tally_success(tally_response):
                    created_tally_voucher = fetch_created_tally_voucher(tally_response)
                    if not created_tally_voucher.get("tallyInvNo"):
                        created_tally_voucher["tallyInvNo"] = invoice_no
                    if not (
                        created_tally_voucher.get("tallyInvDate")
                        and created_tally_voucher.get("tallyInvId")
                    ):
                        voucher_by_number = fetch_tally_voucher_reference(
                            invoice_no,
                            VOUCHER_TYPE_NAME,
                            invoice_row.get("date"),
                            company_name,
                        )
                        if voucher_by_number:
                            created_tally_voucher = {
                                **created_tally_voucher,
                                **voucher_by_number,
                            }
                            if not created_tally_voucher.get("tallyInvNo"):
                                created_tally_voucher["tallyInvNo"] = invoice_no

                    if not (
                        created_tally_voucher.get("tallyInvDate")
                        and created_tally_voucher.get("tallyInvId")
                    ):
                        context_voucher = fetch_tally_voucher_by_context(
                            invoice_row,
                            company_name,
                            narration_text,
                        )
                        if context_voucher:
                            created_tally_voucher = {
                                **created_tally_voucher,
                                **context_voucher,
                            }
                            if not created_tally_voucher.get("tallyInvNo"):
                                created_tally_voucher["tallyInvNo"] = invoice_no

                    has_voucher_no = bool(created_tally_voucher.get("tallyInvNo"))
                    has_voucher_date = bool(created_tally_voucher.get("tallyInvDate"))
                    has_voucher_id = bool(created_tally_voucher.get("tallyInvId"))
                    if has_voucher_no and has_voucher_date and has_voucher_id:
                        remark = "Posted successfully to Tally"
                    elif has_voucher_no:
                        remark = "Posted to Tally; voucher number saved, date/id fetch incomplete"
                    else:
                        remark = "Posted to Tally but voucher number unavailable"
                    update_invoice_tally_status(
                        conn,
                        invoice_id,
                        True,
                        remark,
                        invoice_row.get("updatedBy") or DEFAULT_UPDATED_BY,
                        created_tally_voucher.get("tallyInvNo"),
                        format_iso_date(created_tally_voucher.get("tallyInvDate")),
                        created_tally_voucher.get("tallyInvId"),
                        invoice_row=invoice_row,
                        company_name=company_name,
                        stage="SYNC",
                    )
                    if created_tally_voucher:
                        print(f"Fetched Tally invoice reference: {created_tally_voucher}")
                    print(remark)
                else:
                    error_detail = extract_tally_error(tally_response)
                    update_invoice_tally_status(
                        conn,
                        invoice_id,
                        False,
                        error_detail,
                        invoice_row=invoice_row,
                        company_name=company_name,
                        stage="SYNC",
                    )
                    print(f"Posting failed: {error_detail}")
                    print(
                        f"Stopping batch at Invoice No: {invoice_no}. "
                        "No later invoices will be posted in this run."
                    )
                    break

            except Exception as exc:
                error_message = str(exc)[:1000]
                update_invoice_tally_status(
                    conn,
                    invoice_id,
                    False,
                    error_message,
                    invoice_row=invoice_row,
                    company_name=company_name,
                    stage="SYNC",
                )
                print(f"Error in Invoice ID {invoice_id}: {error_message}")
                print(
                    f"Stopping batch at Invoice No: {invoice_no}. "
                    "No later invoices will be posted in this run."
                )
                break

    finally:
        if conn:
            conn.close()


if __name__ == "__main__":
    try:
        sync_invoices_to_tally()
    except mysql.connector.Error as exc:
        print("==========================================")
        print("Database connection failed")
        print("==========================================")
        print(
            f"Could not connect to MySQL at "
            f"{DB_CONFIG['host']}:{DB_CONFIG['port']}."
        )
        print(f"Connector message: {exc}")
        print(
            "Please check whether outbound access to MySQL port 3306 is "
            "allowed from this machine, and whether the server is accepting "
            "connections from your current IP."
        )
        pause_before_exit()
        sys.exit(1)
    except Exception as exc:
        print(f"Unexpected error: {exc}")
        traceback.print_exc()
        pause_before_exit()
        sys.exit(1)
